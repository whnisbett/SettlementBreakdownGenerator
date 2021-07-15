import pandas as pd
import regex
import itertools
from pathlib import Path
import numpy as np
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.numbers import BUILTIN_FORMATS
from tkinter import Tk, messagebox
from tkinter.filedialog import askopenfilenames

class CloseOutStatement:
    """
    Object containing close out statement information pertaining to a settlement.
    
    ASSUMPTIONS:
    ------------
    - Item names always reside in column 1 (Col B)
    - Numerical values for each item resides in column 9 (Col J) of the corresponding row
    - Sub-total labels reside in column 2 (Col C)
    - Only 3 sections in closeout statement in following order: Settlement info, Expense, Medical/Lien
    - Net to client is a row (and is unique)
    - Total expenses is a row (and is unique
    - Total medical is a row (and is unique)
    - "Amount of Settlement" is a row (and is unique
    """
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.closeout_df = self.import_closeout_as_df(file_path)
        self.preprocess_closeout_df()

    def import_closeout_as_df(self, file_path: str):
        """
        Import closeout statement excel file as a pandas DataFrame
        """
        df = pd.read_excel(file_path, engine='openpyxl', header=None)
        return df

    def preprocess_closeout_df(self):
        """
        Preprocess closeout_df by performing the following operations:
            1. Drop unessential columns
            2. Combine separate item columns together (to get sub-total items into same column as other items)
            3. Rename columns
            4. Remove empty rows
            5. Format all strings
        """
        self._drop_columns_closeout()
        self._combine_item_columns_closeout()
        self._rename_columns_closeout()
        self._drop_empty_rows_closeout()
        self._drop_verbose_rows_closeout()
        self._format_item_entries_closeout()
    
    def _drop_columns_closeout(self):
        """
        Remove unessential columns from closeout_df 
        """
        self.closeout_df = self.closeout_df[[1,2,9]]

    def _combine_item_columns_closeout(self):
        """
        Combine separate item columns together (to get sub-total items into same column as other items)
        """
        self.closeout_df = self.closeout_df.combine_first(self.closeout_df[[2]].rename(columns={2: 1})).drop(columns=[2])
    
    def _rename_columns_closeout(self):
        """
        Rename columns of closeout_df to "item" and "amount"
        """
        self.closeout_df = self.closeout_df.rename(columns={1: 'item', 9: 'amount'})

    def _drop_empty_rows_closeout(self):
        """
        Drop empty rows from closeout_df and reset the indices to be contiguous
        """
        self.closeout_df = self.closeout_df.dropna(how='all', axis=0).reset_index(drop=True)

    def _drop_verbose_rows_closeout(self, len_thresh = 150):
        """
        Drop rows where item column is verbose (> 150 characters)
        """
        len_mask = (self.closeout_df.astype(str).applymap(len) > len_thresh).any(axis=1)
        self.closeout_df = self.closeout_df[~len_mask].reset_index(drop=True)

    def _format_item_entries_closeout(self):
        """
        Format entries in closeout_df
        """
        self.closeout_df['item'] = self.closeout_df['item'].apply(self._format_items)
    
    def _format_items(self, item: str):
        """
        Method for formatting a single string entry in "item" column of closeout_df. Meant to be used with pandas.DataFrame.apply
        """
        item = item.lower()
        item = item.replace(":", '')
        item = item.replace("-", '')
        item = item.strip()
        return item

    def parse_closeout_df(self):
        """
        Parse information in closeout_df and store key information as attributes
        """
        self.client_name = self.get_client_name()
        self.settlement_amount = self.get_settlement_amount()
        self.net_to_client_amount = self.get_net_to_client_amount()
        self.total_expenses_amount = self.get_total_expenses()
        self.total_medical_amount = self.get_total_medical()
        self.medical_items = self.get_medical_items()


    def split_closeout_df(self):
        """
        Split the closeout df into 3 dfs based on the items items each section contains: settlement info, expenses, and medical
        """
        subtotal_mask = self._fuzzy_match_series(self.closeout_df['item'], 'subtotal', errors=2)
        subtotal_idxs = self.closeout_df[subtotal_mask].index
        
        settlement_df = self.closeout_df.iloc[:subtotal_idxs[0] + 1].reset_index(drop=True)
        expenses_df = self.closeout_df.iloc[subtotal_idxs[0] + 1: subtotal_idxs[1] + 1].reset_index(drop=True)
        medical_df = self.closeout_df.iloc[subtotal_idxs[1] + 1:subtotal_idxs[2] + 1].reset_index(drop=True)

        return settlement_df, expenses_df, medical_df

    def get_medical_items(self):
        """
        Get itemized medical expenses from closeout_df
        """
        _, _, medical_df= self.split_closeout_df()
        medical_items = medical_df.iloc[0:-3].dropna().reset_index(drop=True)
        return medical_items
    
    def get_settlement_amount(self):
        """
        Get settlement amount from closeout_df
        """
        settlement_amount_mask = self._fuzzy_match_series(self.closeout_df['item'], 'amount of settlement', errors=3)
        settlement_amount = self.closeout_df[settlement_amount_mask]['amount'].values[0]
        return settlement_amount

    def get_net_to_client_amount(self):
        """
        Get net amount to client from closeout_df
        """
        net_to_client_mask = self._fuzzy_match_series(self.closeout_df['item'], 'net to client', errors=3)
        net_to_client_amount = self.closeout_df[net_to_client_mask]['amount'].values[0]
        return net_to_client_amount

    def get_total_expenses(self):
        """
        Get total amount of all expenses in closeout_df
        """
        total_expenses_mask = self._fuzzy_match_series(self.closeout_df['item'], 'total expenses', errors=3)
        total_expenses = self.closeout_df[total_expenses_mask]['amount'].values[0]
        return total_expenses

    def get_total_medical(self):
        """
        Get total medical medical expenses from closeout_df
        """
        total_medical_mask = self._fuzzy_match_series(self.closeout_df['item'], 'total medical', errors=3)
        total_medical = self.closeout_df[total_medical_mask]['amount'].values[0]
        return total_medical

    def get_client_name(self):
        """
        Get client name from closeout_df
        """
        client_name_mask = self._fuzzy_match_series(self.closeout_df['item'], 'name', errors=1)
        name = self.closeout_df[client_name_mask]['item'].values[0]
        name = name.replace('name', '').strip()
        return name

    def _fuzzy_match_series(self, series: pd.Series, match: str, errors=3):
        """
        Find elements in series that are similar to pattern within some number of errors. Output is a mask of shape series.shape
        """
        series = series.astype(str)
        mask = series.apply(lambda item: self._is_fuzzy_match(item, match=match, errors=errors))
        return mask
        
    def _is_fuzzy_match(self, item: str, match: str, errors=3):
        """
        Returns whether item matches pattern within some number of errors
        """ 
        regex_pattern = f"({match}){{e<={errors}}}"
        matches = regex.findall(regex_pattern, item)
        return len(matches) > 0


class BreakdownWriter:
    """
    Class for transcribing CloseOutStatement objects as breakdown sheets
    """
    def __init__(self, closeout_statement: CloseOutStatement, is_lit=False):
        self.closeout_statement = closeout_statement
        self.is_lit = is_lit
        self.title_final_row = 8
        self.med_table_final_row = self.get_med_table_final_row()
        self.med_table_rows = list(range(self.title_final_row + 1, self.title_final_row + 1 + len(self.closeout_statement.get_medical_items()) + 1))
        self.bottom_section_first_row = self.med_table_final_row + 2
        self.rates_top_row = 34

        self.workbook = self.initialize_workbook()
        self.sheet = self.workbook.active
        self.format_cell_dimensions()
        self.create_title()
        self.insert_column_a_financials()
        self.insert_medical_table()
        self.insert_column_b_financials()
        self.insert_payouts_section()
        self.insert_rates()

    def get_med_table_final_row(self):
        """
        Get the number of the medical services table's final row. 

        FORMULA
        -------
        initial_row + number of medical items + 1 extra row for total = 9 + n_items + 1 = 10 + n_items
        """
        return 10 + len(self.closeout_statement.get_medical_items())

    def initialize_workbook(self):
        """
        Initialize Excel workbook for writing
        """
        workbook = Workbook()
        return workbook

    def format_cell_dimensions(self):
        """
        Format row and column dimensions across entire worksheet
        """
        for row in range(9,40):
            self.sheet.row_dimensions[row].height = 30
            
        self.sheet.column_dimensions['A'].width = 25.17
        self.sheet.column_dimensions['B'].width = 40
        self.sheet.column_dimensions['C'].width = 15
        self.sheet.column_dimensions['D'].width = 22
        self.sheet.column_dimensions['E'].width = 15
        self.sheet.column_dimensions['F'].width = 25
        self.sheet.column_dimensions['G'].width = 20

    def create_title(self):
        """
        Create title at top of statement 
        """
        for row in range(1,self.title_final_row + 1):
            self.sheet.merge_cells(f'B{row}:F{row}')
            self.sheet[f'B{row}'].fill = PatternFill("solid", fgColor="FFFFFF")
        self.sheet['B5'] = 'STATEMENT OF BREAKDOWN'
        self.sheet['B6'] = self.closeout_statement.get_client_name()
        
        self.sheet['B5'].font = Font(color='366092', name='Arial (Headings)', size=12, bold=True)
        self.sheet['B5'].alignment = Alignment(horizontal='center', vertical='center')
        
        self.sheet['B6'].font = Font(color='366092', name='Arial (Body)', size=11)
        self.sheet['B6'].alignment = Alignment(horizontal='center', vertical='center')

    def insert_column_a_financials(self):
        """
        Insert headers and values for each of the financials reported in column A
        """
        self.insert_column_a_financials_headers()
        self.insert_column_a_financials_amounts()

    def insert_column_a_financials_headers(self):
        """
        Insert headers for each of the financials reported in column A
        """
        header_values = ['Settlement Amount',
            'Amount Stated to Providers',
            '1/3 of Settlement for Meds',
            'Attorney Fees',
            '% to Business',
            'OJ',
            'Expenses',
            'Total to Business', 
            'Net to Client']
        header_rows = range(self.title_final_row + 1, self.title_final_row + 1 + len(header_values) * 2, 2)

        for row, val in zip(header_rows, header_values):
            cell = f'A{row}'
            self.sheet[cell] = val
            self.format_cell_as_header(cell)

    def format_cell_as_header(self, cell):
        """
        Format a cell as a blue header
        """
        self.sheet[cell].fill = PatternFill("solid", fgColor="366092")
        self.sheet[cell].font = Font(color="FFFFFF", name='Arial (Body)', size=11, bold=True)
        self.sheet[cell].alignment = Alignment(horizontal='center', vertical='bottom')

    def insert_column_a_financials_amounts(self):
        """
        Insert values for each of the financials reported in column A
        """

        attorney_fees_factor = 0.4 if self.is_lit else 1/3
        amount_values = [self.closeout_statement.get_settlement_amount(),
                        0,
                        '=A10/3',
                        f'=(A10*{attorney_fees_factor})-B{self.bottom_section_first_row + 5}',
                        f'=A16*B{self.rates_top_row + 1}',
                        f'=A14-D{self.med_table_final_row}',
                        abs(self.closeout_statement.get_total_expenses()),
                        '=A18+A20+A22',
                        f'=A10-A16-A20-A22-D{self.med_table_final_row}-B{self.bottom_section_first_row + 3}']
        amount_rows = range(self.title_final_row + 2, self.title_final_row + 2 + len(amount_values) * 2, 2)

        for row, val in zip(amount_rows, amount_values):
            cell = f'A{row}'
            self.sheet[cell] = val
            self.format_cell_as_dollar(cell)

    def format_cell_as_dollar(self, cell):
        """
        Format a cell as a dollar amount
        """
        self.sheet[cell].font = Font(color="000000", name='Arial (Body)', size=11, bold=False)
        self.sheet[cell].alignment = Alignment(horizontal='left', vertical='bottom')
        self.sheet[cell].number_format = BUILTIN_FORMATS[44]

    def insert_medical_table(self):
        """
        Insert medical expenses into the sheet as a formatted table
        """
        med_table = self.generate_medical_table_items()
        med_table_flat = [item  for col in list(zip(*med_table)) for item in col]

        med_cells = self.generate_medical_table_cells()
        med_cells = [''.join(cell) for cell in med_cells]

        for cell, val in zip(med_cells, med_table_flat):
            self.sheet[cell] = val

        self.insert_total_medical_row()
        self.format_medical_items_as_table()
        self.format_medical_table()

    def generate_medical_table_items(self):
        """
        Generate medical expense items from closeout statement in a table format
        """
        med_items = self.closeout_statement.get_medical_items()
        med_items['amount'] = med_items['amount'].abs()
        med_items['Paid/Projected Amount'] = 0
        med_items['Comments'] = '-'
        med_items.values.tolist()
        med_headers = ['Medical Provider', 'Billed Amount', 'Paid/Projected Amount', 'Comments']
        med_table = [med_headers] + med_items.values.tolist()
        return med_table
    
    def generate_medical_table_cells(self):
        """
        Generate corresponding cells for each item in the medical expense list
        """
        med_cols = ['B', 'C', 'D', 'E']
        med_rows = [str(row) for row in self.med_table_rows]
        med_cells = list(itertools.product(med_cols, med_rows))
        return med_cells

    def insert_total_medical_row(self):
        """
        Insert a row for total medical expenses in last row of table
        """
        self.sheet[f'B{self.med_table_final_row}'] = 'Total'
        self.sheet[f'C{self.med_table_final_row}'] = f'=SUM(C{self.med_table_rows[1]}:C{self.med_table_rows[-1]})'
        self.sheet[f'D{self.med_table_final_row}'] = f'=SUM(D{self.med_table_rows[1]}:D{self.med_table_rows[-1]})'

    def format_medical_items_as_table(self):
        """
        Format items in sheet as an Excel table
        """
        tab = Table(displayName="MedicalItems", ref=f"B{self.title_final_row + 1}:E{self.med_table_final_row}")
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        self.sheet.add_table(tab)

    def format_medical_table(self):
        """
        Format the medical items table
        """
        self.format_medical_table_headers()
        self.format_medical_table_values()

    def format_medical_table_headers(self):
        """
        Format headers of the medical items table
        """
        med_cols = ['B', 'C', 'D', 'E']
        for col in med_cols:
            cell = f'{col}{self.title_final_row + 1}'
            self.format_cell_as_header(cell)
        
    def format_medical_table_values(self):
        """
        Format values of the medical items table
        """
        for row in self.med_table_rows[1:] + [self.med_table_final_row]:
            b_cell = f'B{row}'
            c_cell = f'C{row}'
            d_cell = f'D{row}'
            e_cell = f'E{row}'
            
            self.format_cell_as_text(b_cell, bold=True)
            self.format_cell_as_dollar(c_cell)
            self.format_cell_as_dollar(d_cell)
            self.format_cell_as_text(e_cell, center=True)
    
    def format_cell_as_text(self, cell, center=False, bold=False, size=11, color="000000"):
        """
        Format a cell as plain text with the options to change center, bold, size, and color properties
        """
        self.sheet[cell].font = Font(color=color, name='Arial (Body)', size=size, bold=bold)
        if center:
            self.sheet[cell].alignment = Alignment(horizontal='center', vertical='bottom')

    def insert_column_b_financials(self):
        """
        Insert headers and values for each of the financials reported in column B
        """
        self.insert_column_b_financials_headers()
        self.insert_column_b_financials_amounts()

    def insert_column_b_financials_headers(self):
        """
        Insert headers for each of the financials reported in column B
        """
        n_financial_items = 5
        header_rows = range(self.bottom_section_first_row, self.bottom_section_first_row + n_financial_items * 2, 2)
        header_values = ['Total to be Split to Owners',
        '33.33% OR 10% OR Joe Ref INSERT AMOUNT BELOW',
        'Google Atty Portion',
        'Google Business Portion',
        'INSERT 500 BELOW IF LINDALE (goes to mkt acct)']

        for row, val in zip(header_rows, header_values):
            cell = f'B{row}'
            self.sheet[cell] = val
            self.format_cell_as_header(cell)
    
    def insert_column_b_financials_amounts(self):
        """
        Insert values for each of the financials reported in column B
        """
        n_financial_items = 5
        amount_rows = range(self.bottom_section_first_row + 1, self.bottom_section_first_row + n_financial_items * 2, 2)
        amount_values = [f'=(A16*B{self.rates_top_row}) - B{self.bottom_section_first_row + 5}',
                0,
                0,
                0,
                0]

        for row, val in zip(amount_rows, amount_values):
            cell = f'B{row}'
            self.sheet[cell] = val
            self.format_cell_as_dollar(cell)
    
    def insert_payouts_section(self):
        """
        Insert payout totals for each account into sheet
        """
        n_accounts = 10
        account_rows = range(self.bottom_section_first_row, self.bottom_section_first_row + n_accounts)
        account_names = ['AP/CN',
                        'Jerry',
                        'CASE MANAGER',
                        'Business',
                        'Total to Business',
                        'MKT',
                        'CSH',
                        '10% or 1/3 or Joe Referral (Write From Operating and give 1099)',
                        'BLD']

        account_amounts = [f'=(B{self.bottom_section_first_row + 1}-D{self.bottom_section_first_row + 1})/2',
                        f'=B{self.bottom_section_first_row + 1}*0.13',
                        f'=A16*B{self.rates_top_row + 2}',
                        f'=(A24-B{self.bottom_section_first_row + 7})',
                        f'=D{self.bottom_section_first_row}*2 + D{self.bottom_section_first_row + 1} + D{self.bottom_section_first_row + 2} + D{self.bottom_section_first_row + 3} + D{self.bottom_section_first_row + 6} + B{self.bottom_section_first_row + 3} + B{self.bottom_section_first_row + 5} + B{self.bottom_section_first_row + 7}',
                        f'=A16*0.2825+(B{self.bottom_section_first_row + 5}+B{self.bottom_section_first_row + 7}+B{self.bottom_section_first_row + 9})',
                        f'=A16*B{self.rates_top_row + 3}',
                        f'=B{self.bottom_section_first_row + 3}',
                        '=A16*0.04']

        for row, name, amount in zip(account_rows, account_names, account_amounts):
            self.sheet[f'C{row}'] = name
            self.sheet[f'D{row}'] = amount
            self.format_cell_as_dollar(f'D{row}')

        self.sheet[f'F{self.bottom_section_first_row}'] = 'Total to Operating'
        self.sheet[f'G{self.bottom_section_first_row}'] = f'=D{self.bottom_section_first_row + 4} - SUM(D{self.bottom_section_first_row}:D{self.bottom_section_first_row + 2}) - SUM(D{self.bottom_section_first_row + 5}:D{self.bottom_section_first_row + 8}) - D{self.bottom_section_first_row} - A20'

        self.format_payouts_section()
    
    def format_payouts_section(self):
        """
        Format accounts total section
        """
        rows = [self.bottom_section_first_row + i for i in [4, 5, 6, 8]]
        colors = ['C4D79B', 'F79646', 'B1A0C7', '4BACC6']

        for row, color in zip(rows, colors):
            self.sheet[f'D{row}'].fill = PatternFill("solid", fgColor=color)
            self.sheet[f'E{row}'].fill = PatternFill("solid", fgColor=color)
            self.sheet[f'C{row}'].fill = PatternFill("solid", fgColor=color)

        self.sheet[f'F{self.bottom_section_first_row}'].fill = PatternFill("solid", fgColor="C4BD97")
        self.sheet[f'G{self.bottom_section_first_row}'].fill = PatternFill("solid", fgColor="C4BD97")
        self.format_cell_as_dollar(f'G{self.bottom_section_first_row}')

    def insert_rates(self):
        """
        Insert rates section at bottom of sheet
        """
        rate_names = ['AP/CN',
                      'Firm',
                      'Commission',
                      'To Cash\'s',
                      'THIS CELL MUST ALWAYS EQUAL 1.0000',
                      'Jerry % Calculation']
        rate_amounts = [0.25,
                        0.71,
                        0.01,
                        0.07 - 0.04,
                        f'=SUM(B{self.rates_top_row}:B{self.rates_top_row + 3})',
                        f'=0.06*(1-B{self.rates_top_row + 3})']
        rate_rows = range(self.rates_top_row,self.rates_top_row + len(rate_names))

        for row, name, amount in zip(rate_rows, rate_names, rate_amounts):
            self.sheet[f'A{row}'] = name
            self.sheet[f'B{row}'] = amount

        self.insert_rates_comments()
    
    def insert_rates_comments(self):
        """
        Insert comments about rates into sheet
        """
        self.sheet['C34'] = '<- DO NOT CHANGE'
        self.sheet['C35'] = '<- DO NOT CHANGE'
        self.sheet['C39'] = 'CHANGE 0.06 remains constant.'

        self.sheet['D35'] = 'From 0.71 change to 0.705'
        self.sheet['D36'] = 'For 1.5% change to 0.015'

        self.sheet['E35'] = 'From 0.71 change to 0.69'
        self.sheet['E36'] = 'For 3% (Ana + Ray) change to 0.03'

    def save_workbook(self, output_file):
        """
        Write workbook to filepath
        """
        self.workbook.save(output_file)
        # self._set_password(output_file, 'cindyflorist')

    def _set_password(self, excel_file_path, pw):

        import subprocess

        excel_file_path = Path(excel_file_path)

        vbs_script = \
        f"""' Save with password required upon opening

        Set excel_object = CreateObject("Excel.Application")
        Set workbook = excel_object.Workbooks.Open("{excel_file_path}")

        excel_object.DisplayAlerts = False
        excel_object.Visible = False

        workbook.SaveAs "{excel_file_path}",, "{pw}"

        excel_object.Application.Quit
        """

        # write
        vbs_script_path = excel_file_path.parent.joinpath("set_pw.vbs")
        with open(vbs_script_path, "w") as file:
            file.write(vbs_script)

        #execute
        subprocess.call(['cscript.exe', str(vbs_script_path)])

        # remove
        vbs_script_path.unlink()

        return None


if __name__ == "__main__":
    Tk().withdraw()
    filenames = askopenfilenames()
    statements_and_parents = [(CloseOutStatement(path), Path(path).absolute()) for path in filenames]
    for statement, path in statements_and_parents:
        parent_path = path.parent
        client_name = statement.get_client_name().upper()
        output_file = parent_path / f'Breakdown {client_name}.xlsx'
        import pdb
        pdb.set_trace()
        response = messagebox.askquestion("Litigation?", f"File: {path.stem} \n\n Has this gone to litigation?", icon='question')
        is_lit = True if response == 'yes' else False
        writer = BreakdownWriter(statement, is_lit=is_lit)
        writer.save_workbook(output_file)
